import pandas as pd
import os
import base64
from pathlib import Path
from typing import List, Dict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config.settings import settings
from utils.helpers import generate_report_filename, format_datetime
from utils.logger import get_logger
# 👇 1. IMPORTAR SERVICIO DE DRIVE
from services.google.drive_service import GoogleDriveService

logger = get_logger(__name__)

class ExcelReportGenerator:
    '''Generador optimizado de reportes en Excel con formato profesional y PDFs adjuntos'''
    
    def __init__(self):
        self.reports_dir = settings.REPORTS_DIR
        self.pdf_dir = self.reports_dir / 'pdf' # Carpeta temporal
        # 2. INICIALIZAR SERVICIO
        self.drive_service = GoogleDriveService()
        
        # Asegurar carpeta temporal
        self.pdf_dir.mkdir(parents=True, exist_ok=True)
    
    def generar_reporte(self, resultados: List[Dict]) -> str:
        '''Genera un archivo Excel profesional con los resultados y links a PDFs'''
        logger.info('📊 Generando reporte Excel profesional...')
        
        # 👇 3. PROCESAR PDFs ANTES DE CREAR EL EXCEL
        # Esto sube los archivos y agrega el campo 'drive_link' a los resultados
        resultados_con_links = self._procesar_pdfs_y_subir(resultados)
        
        # Preparar datos usando los resultados ya enriquecidos con links
        data = self._preparar_datos(resultados_con_links)
        df = pd.DataFrame(data)
        
        # Generar nombre y path
        filename = generate_report_filename()
        filepath = self.reports_dir / filename
        
        # Crear Excel con formato
        self._crear_excel_formateado(df, filepath, resultados)
        
        logger.info(f'✅ Reporte generado: {filepath.name}')
        return str(filepath)

    # 👇 4. NUEVO MÉTODO: REHIDRATA PDF, SUBE A DRIVE Y BORRA LOCAL
    def _procesar_pdfs_y_subir(self, resultados: List[Dict]) -> List[Dict]:
        '''Itera sobre resultados, rehidrata PDF, sube a Drive, obtiene Link y borra local'''
        
        for res in resultados:
            res['drive_link'] = "N/A"
            
            # Obtener base64 crudo (pasado desde el validador)
            base64_str = res.get('factura_base64_raw')
            
            if base64_str and len(base64_str) > 100:
                try:
                    # Limpiar header de data uri si existe
                    if ',' in base64_str:
                        base64_str = base64_str.split(',')[1]
                        
                    pdf_bytes = base64.b64decode(base64_str)
                    
                    # Nombre único para el PDF temporal
                    pdf_name = f"Factura_{res['registro_id']}_{format_datetime(res['fecha_recepcion'], '%Y%m%d_%H%M%S')}.pdf"
                    pdf_path = self.pdf_dir / pdf_name
                    
                    # Guardar localmente
                    with open(pdf_path, 'wb') as f:
                        f.write(pdf_bytes)
                    
                    # Subir a Drive (Carpeta Específica de Facturas)
                    link = self.drive_service.upload_to_drive(
                        str(pdf_path), 
                        specific_folder_id=settings.GOOGLE_DRIVE_INVOICES_FOLDER_ID
                    )
                    
                    res['drive_link'] = link
                    
                    # Borrar archivo local (Limpieza)
                    if pdf_path.exists():
                        os.remove(pdf_path)
                    
                except Exception as e:
                    logger.error(f"  ❌ Error procesando PDF registro {res['registro_id']}: {e}")
                    res['drive_link'] = "Error al subir"
            
        return resultados
    
    def _preparar_datos(self, resultados: List[Dict]) -> List[Dict]:
        '''Prepara los datos para el DataFrame con información completa'''
        data = []
        for res in resultados:
            # Determinar estado visual
            if res['es_anomalia']:
                estado = '⚠️ ANOMALÍA'
            else:
                estado = '✅ OK'
            
            fila = {
                'Estado': estado,
                'ID Registro': res['registro_id'],
                'Fecha Recepción': format_datetime(res['fecha_recepcion']),
                'Tiene OC': 'Sí' if res['tiene_oc'] else 'No',
                'Cantidad OCs': res.get('num_ordenes', 0),
                'Tiene Factura': 'Sí' if res['tiene_factura'] else 'No',
                # 👇 5. AGREGAR COLUMNA DE LINK
                'Link Factura': res.get('drive_link', 'N/A'),
                'Tipo Anomalía': self._format_tipo_anomalia(res.get('tipo_anomalia')),
            }
            
            # Agregar datos de factura si existen
            if res.get('datos_factura'):
                df_datos = res['datos_factura']
                fila.update({
                    'Proveedor': df_datos.get('proveedor', 'N/A'),
                    'Gran Total': df_datos.get('gran_total', 'N/A'),
                    'Moneda': df_datos.get('moneda', 'N/A'),
                    'Receptor': df_datos.get('receptor', 'N/A'),
                    'Folio': df_datos.get('folio', 'N/A'),
                })
            else:
                fila.update({
                    'Proveedor': 'N/A',
                    'Gran Total': 'N/A',
                    'Moneda': 'N/A',
                    'Receptor': 'N/A',
                    'Folio': 'N/A',
                })
            
            # Resultado de validación
            validacion = res.get('resultado_openai', 'N/A')
            if validacion and len(validacion) > 200:
                validacion = validacion[:200] + '...'
            fila['Resultado Validación'] = validacion
            
            data.append(fila)
        
        return data
    
    def _format_tipo_anomalia(self, tipo: str) -> str:
        '''Formatea el tipo de anomalía para mejor lectura'''
        if not tipo: return 'N/A'
        formatos = {
            'sin_oc': '⚠️ Sin OC',
            'sin_factura': '⚠️ Sin Factura',
            'sin_oc_ni_factura': '🔴 Sin OC ni Factura',
            'error_procesamiento_factura': '❌ Error al Procesar',
            'discrepancias_encontradas': '⚠️ Discrepancias'
        }
        return formatos.get(tipo, tipo.replace('_', ' ').title())
    
    def _crear_excel_formateado(self, df: pd.DataFrame, filepath: Path, resultados: List[Dict]):
        '''Crea el Excel con formato profesional avanzado'''
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Escribir DataFrame principal
            df.to_excel(writer, sheet_name='Validación', index=False)
            
            # Obtener worksheet
            worksheet = writer.sheets['Validación']
            
            # Aplicar estilos
            self._aplicar_estilos_header(worksheet)
            
            # 👇 6. CONVERTIR LINKS DE TEXTO A HIPERVÍNCULOS REALES
            self._aplicar_hipervinculos(worksheet, df)
            
            self._ajustar_columnas(worksheet)
            self._aplicar_formato_condicional(worksheet, len(df))
            
            # Crear hoja de resumen
            self._crear_hoja_resumen(writer, resultados)
    
    # 👇 MÉTODO NUEVO PARA HIPERVÍNCULOS AZULES
    def _aplicar_hipervinculos(self, worksheet, df):
        # Buscar índice de columna 'Link Factura'
        col_link_idx = None
        for idx, col_name in enumerate(df.columns, 1):
            if col_name == 'Link Factura':
                col_link_idx = idx
                break
        
        if col_link_idx:
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_link_idx)
                val = cell.value
                # Si es un link de Drive válido
                if val and str(val).startswith('http'):
                    cell.value = f'=HYPERLINK("{val}", "Ver PDF")'
                    cell.font = Font(color="0000FF", underline="single")

    def _aplicar_estilos_header(self, worksheet):
        '''Aplica estilos al header'''
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    def _ajustar_columnas(self, worksheet):
        '''Ajusta el ancho de las columnas automáticamente'''
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 12), 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _aplicar_formato_condicional(self, worksheet, num_rows):
        '''Aplica formato condicional según el estado'''
        ok_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        error_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
        
        for row in range(2, num_rows + 2):
            estado_cell = worksheet[f'A{row}']
            
            if '✅' in str(estado_cell.value):
                fill = ok_fill
            elif '⚠️' in str(estado_cell.value):
                fill = warning_fill
            else:
                fill = error_fill
            
            for col in range(1, worksheet.max_column + 1):
                # No sobreescribir el estilo del link si es la columna de link
                cell = worksheet.cell(row=row, column=col)
                # Aplicar fondo a todos, pero mantener fuente azul en links
                if cell.font.color and cell.font.color.rgb == "0000FF":
                    pass # Es un link, no tocamos la fuente
                
                cell.fill = fill

    def _crear_hoja_resumen(self, writer, resultados: List[Dict]):
        '''Crea una hoja de resumen con estadísticas'''
        total = len(resultados)
        anomalias = sum(1 for r in resultados if r['es_anomalia'])
        correctos = total - anomalias
        
        tipos_anomalias = {}
        for r in resultados:
            if r['es_anomalia'] and r.get('tipo_anomalia'):
                tipo = self._format_tipo_anomalia(r['tipo_anomalia'])
                tipos_anomalias[tipo] = tipos_anomalias.get(tipo, 0) + 1
        
        resumen_data = {
            'Métrica': [
                'Total de Registros',
                'Registros Correctos',
                'Total de Anomalías',
                'Porcentaje de Correctos',
                '',
                'DESGLOSE DE ANOMALÍAS'
            ],
            'Valor': [
                total,
                correctos,
                anomalias,
                f'{(correctos/total*100):.1f}%' if total > 0 else '0%',
                '',
                ''
            ]
        }
        
        for tipo, cantidad in sorted(tipos_anomalias.items()):
            resumen_data['Métrica'].append(tipo)
            resumen_data['Valor'].append(cantidad)
        
        df_resumen = pd.DataFrame(resumen_data)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        ws_resumen = writer.sheets['Resumen']
        self._aplicar_estilos_header(ws_resumen)
        ws_resumen.column_dimensions['A'].width = 35
        ws_resumen.column_dimensions['B'].width = 20